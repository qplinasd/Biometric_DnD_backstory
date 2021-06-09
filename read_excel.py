# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import os
import xlsxwriter

if __name__ == '__main__':
    cur_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(cur_dir, "20210503-0855am-PART.xlsx")
    wb = load_workbook(excel_path)

    ws = wb[wb.sheetnames[0]]

    storyBook = xlsxwriter.Workbook('backstory.xlsx')
    ws2 = storyBook.add_worksheet()

    # ID
    ws2.write(0, 0, "ID")
    # sample1: You are an old wizard with your own ideas of life in the alliance of neutral evil.
    # sample2: As an old wizard, you are always misunderstood in the good alignment
    # sample3: All members of the good alignment think that you are an old wizard with your own ideas
    ws2.write(0, 1, "story1")
    # negative personality
    # sample1: People think that you are Aggressive and boring.
    # sample2: Indifference and pride in the eyes makes others dare not close to you.
    # sample3: The suffering of life makes you boring and alert to the closeness of others.
    ws2.write(0, 2, "story2")
    # positive personality
    # sample1: However, everyone who knows you realize that you are caring.
    # sample2: Unlike what it looks like, you're actually calm.
    ws2.write(0, 3, "story3")
    # Birth family
    # sample1: Your parents are always interesting and happy, but ,but you are more introverted.
    # sample2: Living in a happy family atmosphere, you are as intelligent as your parents.
    ws2.write(0, 4, "story4")

    for i in range(1, ws.max_row):
        ws2.write(i, 0, ws["A"][i].value)  # ID

        # story1
        # Q:DD_Race
        # P:DD_Alignment
        # T:BM_Age
        # BC:BM_Kind_Conf

        if ws["T"][i].value < 20:
            age = "a young "
        elif ws["T"][i].value > 60:
            age = "an old "
        else:
            age = "a mature "

        alignment = ws["P"][i].value.lower()

        if ws["BC"][i].value > 0.7 and ('evil' in alignment):
            ws2.write(i, 1, "You are " + age + ws["Q"][i].value.lower()
                      + " with your own ideas of life in the " + alignment +
                      " alignment.")
        elif ws["BC"][i].value <= 0.7 and ('good' in alignment):
            ws2.write(i, 1, "As " + age + ws["Q"][i].value.lower()
                      + ", you are often misunderstood in the " + alignment +
                      " alignment.")
        else:
            ws2.write(i, 1, "All members of the " +
                      alignment + " alignment think that you are "
                      + age + ws["Q"][i].value.lower()
                      + " with your own ideas.")

        # story2
        # W:BM_Aggressive_Conf
        # AA:BM_Boring_Conf
        # AG:BM_Cold_Conf
        # AK:BM_Egotistic_Conf

        negDic = {"aggressive": ws["W"][i].value,
                  "boring": ws["AA"][i].value,
                  "cold": ws["AG"][i].value,
                  "egotistic": ws["AK"][i].value}

        sortedNegDic = sorted(negDic.items(), key=lambda x: x[1])
        if sortedNegDic[0][0] == "aggressive":
            ws2.write(i, 2, "People think that you are "
                      + sortedNegDic[1][0] +
                      " and easy to attack others.")

        elif sortedNegDic[0][0] == "cold" \
                or sortedNegDic[1][0] != "aggressive":

            ws2.write(i, 2, "Indifference and " +
                      sortedNegDic[1][0] +
                      " in the eyes makes others dare not close to you.")
        else:

            ws2.write(i, 2, "The suffering of life makes you " + sortedNegDic[0][0] +
                      " and alert to the closeness of others.")

        # story3
        # Y:BM_Attractive_Conf
        # AC:BM_Calm_Conf
        # AE:BM_Caring_Conf
        # AI:BM_Confident_Conf
        # BC:BM_Kind_Conf
        # AM:BM_Emotinstable_Conf
        # BK:BM_Trustworthy_Conf

        posDic = {"attractive": ws["Y"][i].value,
                  "calm": ws["AC"][i].value,
                  "confident": ws["AI"][i].value,
                  "kind": ws["BC"][i].value,
                  "emotinstable": ws["AM"][i].value
                  }
        sortedPosDic = sorted(posDic.items(), key=lambda x: x[1])

        if ws["BK"][i].value > 0.7:
            ws2.write(i, 3, "However, everyone who knows you realize that you are "
                      + sortedPosDic[0][0] + ".")
        else:
            ws2.write(i, 3, "Unlike what it looks like, you are actually "
                      + sortedPosDic[0][0] + ".")

        # story4
        # AQ:BM_Familiar_Conf
        # AS:BM_Happy_Conf
        # AU:BM_Humble_Conf
        # AW:BM_Intelligent_Conf
        # AY:BM_Interesting_Conf
        # BA:BM_Introverted_Conf

        famDic = {"familiar": ws["AQ"][i].value,
                  "happy": ws["AS"][i].value,
                  "humble": ws["AU"][i].value,
                  "intelligent": ws["AW"][i].value,
                  "interesting": ws["AY"][i].value,
                  "introverted": ws["BA"][i].value
                  }
        sortedFamDic = sorted(famDic.items(), key=lambda x: x[1])

        if sortedFamDic[0][0] == "introverted" or sortedFamDic[1][0] == "introverted":
            ws2.write(i, 4, "Your family are always " +
                      sortedPosDic[2][0] +
                      " and " + sortedPosDic[3][0] + ",but you are more introverted.")
        else:
            ws2.write(i, 4, "Living in a " + sortedFamDic[1][0]
                      + " family atmosphere, you are as "
                      + sortedFamDic[0][0] + " as your parents.")

    storyBook.close()
